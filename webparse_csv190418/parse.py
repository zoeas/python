from urllib.request import Request, urlopen
from bs4 import BeautifulSoup
from openpyxl import Workbook
from collections import OrderedDict
from datetime import datetime
import re
import time

urllist = []
urllist.append('http://www.dgkoaa.or.kr/sub01-08.jsp?common_sel=10&common_code=30') # 중구
urllist.append('http://www.dgkoaa.or.kr/sub01-08.jsp?common_sel=20&common_code=30') # 동구
urllist.append('http://www.dgkoaa.or.kr/sub01-08.jsp?common_sel=30&common_code=30') # 서구
urllist.append('http://www.dgkoaa.or.kr/sub01-08.jsp?common_sel=40&common_code=30') # 남구
urllist.append('http://www.dgkoaa.or.kr/sub01-08.jsp?common_sel=50&common_code=30') # 북구
urllist.append('http://www.dgkoaa.or.kr/sub01-08.jsp?common_sel=60&common_code=30') # 수성구

write_wb = Workbook()

# 엑셀쓰기
write_ws = write_wb.active
write_ws['A1'] = '시작'

html = ''

for url in urllist:
    req = Request(url)
    res = urlopen(req)
    html += res.read().decode('cp949')

bs = BeautifulSoup(html, 'html.parser')
'''
tag_table = bs.find('table', attrs={'class': 'tb03'})
tags_tbody = tag_table.findAll('tbody')
'''

# 모든 td를 찾음
tags_td = bs.findAll('td')

regex = re.compile(r"<td>(\D+.+) (\d+.+) <br\/>")

title_list = []
num_list = []
for tag_td in tags_td:
    rm = tag_td.find('font')
    if not rm :                 # font 태그가 없는 td만
        m = regex.search(str(tag_td))
        if m != None:
            title_list.append(m.group(1).replace(' ',''))
            num_list.append(m.group(2))

# 업체랑 전화번호 저장 (중복제거)
temp_set = set()
rtitle_list = []
rnum_list = []

# 임시집합에 업체명을 저장하면서 리스트 비교, 중복 존재시 패스
for t,n in zip(title_list, num_list):
    if t not in temp_set:
        rtitle_list.append(t)
        rnum_list.append(n)
        temp_set.add(t)

i = 1
for t, n in zip(rtitle_list, rnum_list):   # 1열에 업체명, 2열에 전화번호, 빈행 2개 삽입후 다음 기록 계속
    write_ws.cell(i, 1, t)
    write_ws.cell(i, 2, n)
    i = i + 3

now = datetime.now()
write_wb.save('%s-%s-%s 옥외광고협회 전화번호.xlsx' % (now.year, now.month, now.day))



# tabel 클래스 'tb03' tbody th 패스, 첫번째 tr 패스 tr 3개 연속 첫번째 td 패스, 남은 두개 td br 이전의 택스트만



